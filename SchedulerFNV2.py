#!/usr/bin/env python
# coding: utf-8

from ortools.sat.python import cp_model
from datetime import date, timedelta, datetime
from collections import defaultdict
import json
from openpyxl import Workbook, load_workbook # Aggiunto load_workbook
import math
import os
import glob

# --- Funzioni base ---
def genera_giorni_del_mese(anno, mese):
    """Genera una lista di date (formato stringa) per il mese specificato."""
    giorni = []
    giorno = date(anno, mese, 1)
    while giorno.month == mese:
        giorni.append(giorno.strftime("%Y-%m-%d"))
        giorno += timedelta(days=1)
    return giorni

def get_settimana_mese(data_str):
    """Restituisce il numero della settimana per una data (calendario ISO)."""
    data_obj = date.fromisoformat(data_str)
    return data_obj.isocalendar()[1]


def get_giorno_settimana(data_str):
    """Restituisce il giorno della settimana in italiano."""
    giorni_italiano = {
        0: "Lunedì",
        1: "Martedì",
        2: "Mercoledì",
        3: "Giovedì",
        4: "Venerdì",
        5: "Sabato",
        6: "Domenica"
    }
    data_obj = date.fromisoformat(data_str)
    giorno_idx = data_obj.weekday()
    giorno_it = giorni_italiano[giorno_idx]
    return giorno_it

def get_next_version_number(base_name, extension, directory):
    """Trova il numero di versione successivo per il file Excel."""
    pattern = os.path.join(directory, f"{base_name}_v*")
    existing_files = glob.glob(pattern)
    versions = [int(os.path.basename(f).split('_v')[-1].split('.')[0]) for f in existing_files if f.endswith(extension)]
    next_version = max(versions, default=0) + 1
    return next_version

# --- NUOVA FUNZIONE: Caricamento Indisponibilità da Excel ---
def carica_indisponibilita_da_excel(file_path, nome_foglio='Foglio1'):
    """
    Carica le date di indisponibilità dei dipendenti da un file Excel.

    Args:
        file_path (str): Il percorso del file Excel.
        nome_foglio (str): Il nome del foglio Excel da leggere.

    Returns:
        dict: Un dizionario dove le chiavi sono i nomi dei dipendenti (normalizzati in maiuscolo)
              e i valori sono liste di date (stringhe "YYYY-MM-DD")
              in cui il dipendente non è disponibile.
    """
    indisponibilita = defaultdict(list)
    try:
        workbook = load_workbook(filename=file_path)
        sheet = workbook[nome_foglio]

        # Salta l'intestazione (prima riga)
        rows = iter(sheet.rows)
        try:
            next(rows) # Salta l'intestazione
        except StopIteration:
            print(f"Attenzione: il foglio '{nome_foglio}' nel file '{file_path}' è vuoto o contiene solo l'intestazione.")
            return indisponibilita


        for row_idx, row in enumerate(rows, start=2): # start=2 per considerare la riga di intestazione saltata
            if not row or len(row) < 2: # Controlla se la riga è vuota o non ha abbastanza colonne
                # print(f"Attenzione: Riga {row_idx} vuota o incompleta nel file Excel.")
                continue

            nome_dipendente_excel_cell = row[0]
            data_non_disponibile_excel_cell = row[1]

            if nome_dipendente_excel_cell is None or nome_dipendente_excel_cell.value is None:
                # print(f"Attenzione: Nome dipendente mancante alla riga {row_idx} del file Excel.")
                continue
            if data_non_disponibile_excel_cell is None or data_non_disponibile_excel_cell.value is None:
                # print(f"Attenzione: Data di indisponibilità mancante per {nome_dipendente_excel_cell.value} alla riga {row_idx} del file Excel.")
                continue

            nome_dipendente_excel = nome_dipendente_excel_cell.value
            data_non_disponibile_excel = data_non_disponibile_excel_cell.value

            # Assicurati che il nome del dipendente sia una stringa e normalizzalo (es. maiuscolo)
            nome_dipendente_norm = str(nome_dipendente_excel).strip().upper()

            # Converti la data in formato stringa "YYYY-MM-DD" se necessario
            data_str = ""
            if isinstance(data_non_disponibile_excel, datetime):
                data_str = data_non_disponibile_excel.strftime("%Y-%m-%d")
            elif isinstance(data_non_disponibile_excel, str):
                # Prova a parsare la stringa se è in un formato comune come GG/MM/AAAA
                try:
                    # Tenta prima il formato GG/MM/AAAA
                    data_obj = datetime.strptime(data_non_disponibile_excel, "%d/%m/%Y")
                    data_str = data_obj.strftime("%Y-%m-%d")
                except ValueError:
                    # Se fallisce, prova a vedere se è già YYYY-MM-DD
                    try:
                        datetime.strptime(data_non_disponibile_excel, "%Y-%m-%d") # Solo per validazione
                        data_str = data_non_disponibile_excel
                    except ValueError:
                        print(f"Attenzione: formato data non riconosciuto '{data_non_disponibile_excel}' per {nome_dipendente_norm} alla riga {row_idx}. Usare YYYY-MM-DD o GG/MM/AAAA.")
                        continue # Salta questa riga se il formato data non è gestito
            elif isinstance(data_non_disponibile_excel, (int, float)): # Gestione date Excel come numeri seriali
                 try:
                    # openpyxl di solito converte automaticamente, ma in caso di lettura diretta come numero:
                    # Questa è una conversione semplificata, potrebbe necessitare di aggiustamenti
                    # a seconda del sistema di data di Excel (1900 o 1904)
                    # Per una gestione robusta, assicurarsi che openpyxl interpreti correttamente i tipi di cella.
                    # Se openpyxl restituisce un float, è probabile che sia già stato convertito o è un errore.
                    # Se è un int che rappresenta una data seriale:
                    from openpyxl.utils.datetime import from_excel
                    data_obj = from_excel(data_non_disponibile_excel)
                    data_str = data_obj.strftime("%Y-%m-%d")
                 except Exception as e_date_serial:
                    print(f"Attenzione: formato data numerico non convertibile '{data_non_disponibile_excel}' per {nome_dipendente_norm} alla riga {row_idx}: {e_date_serial}")
                    continue
            else:
                print(f"Attenzione: tipo di dato non riconosciuto per la data '{data_non_disponibile_excel}' ({type(data_non_disponibile_excel)}) per {nome_dipendente_norm} alla riga {row_idx}.")
                continue

            if data_str:
                indisponibilita[nome_dipendente_norm].append(data_str)

        if indisponibilita:
            print(f"Indisponibilità caricate con successo da {file_path}")
            # print(f"Dati indisponibilità caricati: {dict(indisponibilita)}") # Per debug
        else:
            print(f"Nessuna indisponibilità valida trovata in {file_path} nel foglio {nome_foglio}.")
        return indisponibilita
    except FileNotFoundError:
        print(f"Errore: il file di indisponibilità '{file_path}' non è stato trovato.")
        return indisponibilita # Restituisce un dizionario vuoto
    except KeyError:
        print(f"Errore: il foglio '{nome_foglio}' non è stato trovato nel file '{file_path}'. Controlla il nome del foglio.")
        return indisponibilita
    except Exception as e:
        print(f"Errore imprevisto durante la lettura del file Excel '{file_path}': {e}")
        return indisponibilita # Restituisce un dizionario vuoto


# --- Funzioni per la pianificazione ---

def carica_configurazione(config_path):
    """Carica e valida la configurazione dal file JSON."""
    try:
        with open(config_path, "r", encoding="utf-8") as f:
            config = json.load(f)
        print(f"Configurazione caricata con successo da {config_path}")
        return config
    except FileNotFoundError:
        print(f"Errore: il file di configurazione non è stato trovato in {config_path}")
        exit()
    except json.JSONDecodeError as e:
        print(f"Errore nel parsing del file JSON: {e}")
        print(f"Controlla la sintassi del file di configurazione.")
        exit()

# MODIFICATA: inizializza_dati
def inizializza_dati(config, file_indisponibilita_path):
    """Inizializza le strutture dati necessarie dal file di configurazione e indisponibilità."""
    anno_pianificazione = config["anno"]
    mese_pianificazione = config["mese"]
    lista_cantieri_config = config["cantieri"]
    lista_turni_config = config["turni"]
    lista_dipendenti_config = config["dipendenti"]
    vincoli_globali = config.get("vincoli_globali", {})
    dipendenti_per_turno_config = vincoli_globali.get("dipendenti_per_turno", {})

    # Carica le indisponibilità da Excel
    # Assumiamo che il file Excel sia nella stessa directory dello script se non specificato diversamente
    # Il nome del file è passato come argomento
    indisponibilita_dipendenti_excel = carica_indisponibilita_da_excel(file_indisponibilita_path)


    giorni_mese = genera_giorni_del_mese(anno_pianificazione, mese_pianificazione)
    num_giorni = len(giorni_mese)
    num_cantieri = len(lista_cantieri_config)
    nome_cantieri = {i: c["nome"] for i, c in enumerate(lista_cantieri_config)}
    turni_per_cantiere = {i: c["turni_disponibili"] for i, c in enumerate(lista_cantieri_config)}
    nome_turni_map = {t["nome"]: i for i, t in enumerate(lista_turni_config)}
    indice_turni_map = {i: t["nome"] for i, t in enumerate(lista_turni_config)}
    durata_turni = {t["nome"]: int(t["durata_ore"] * 100) for t in lista_turni_config}
    orario_inizio_turni = {t["nome"]: t.get("orario_inizio", 0) for t in lista_turni_config}
    certificazioni_richieste = {t["nome"]: t.get("richiedi_certificazione") for t in lista_turni_config}
    quando_turni = {t["nome"]: t["quando"] for t in lista_turni_config}

    dipendenti = []
    # Normalizza i nomi dei dipendenti dalla configurazione per il confronto
    nome_dipendente = {d["id"]: d["nome"].strip().upper() for d in lista_dipendenti_config}
    id_a_indice_dipendente = {d["id"]: i for i, d in enumerate(lista_dipendenti_config)}

    for dip_config in lista_dipendenti_config:
        dipendenti.append({**dip_config})
    num_dipendenti = len(dipendenti)

    print("Dati inizializzati.")
    return {
        "giorni_mese": giorni_mese,
        "num_giorni": num_giorni,
        "num_cantieri": num_cantieri,
        "nome_cantieri": nome_cantieri,
        "turni_per_cantiere": turni_per_cantiere,
        "nome_turni_map": nome_turni_map,
        "indice_turni_map": indice_turni_map,
        "durata_turni": durata_turni,
        "orario_inizio_turni": orario_inizio_turni,
        "certificazioni_richieste": certificazioni_richieste,
        "quando_turni": quando_turni,
        "dipendenti": dipendenti,
        "nome_dipendente": nome_dipendente, # Nomi già normalizzati
        "id_a_indice_dipendente": id_a_indice_dipendente,
        "num_dipendenti": num_dipendenti,
        "dipendenti_per_turno_config": dipendenti_per_turno_config,
        "vincoli_globali": vincoli_globali,
        "indisponibilita_dipendenti_excel": indisponibilita_dipendenti_excel # Aggiunto questo
    }

def calcola_priorita_vincoli(dipendente):
    """Calcola un punteggio di priorità per il dipendente basato sui vincoli."""
    vincoli = dipendente.get("vincoli", [])
    punteggio = 0
    for vincolo in vincoli:
        if vincolo.get("tipo") not in ["limite_ore_settimanali", "alternanza_domenicale_globale"]:
            if not vincolo.get("soft", False):
                punteggio += 2
            else:
                punteggio += 1
            tipo = vincolo.get("tipo")
            if tipo in ["esclusione_temporale", "esclusione_oraria", "restrizione_cantieri"]:
                 punteggio += 1
    return punteggio

def crea_variabili_modello(model, dati):
    """Crea le variabili di assegnamento e deficit per il modello CP-SAT."""
    assegnamenti = {}
    for d_idx in range(dati["num_dipendenti"]):
        for g_idx in range(dati["num_giorni"]):
            for c_idx in range(dati["num_cantieri"]):
                for nome_turno_loop in dati["turni_per_cantiere"][c_idx]:
                    t_idx = dati["nome_turni_map"][nome_turno_loop]
                    assegnamenti[(d_idx, g_idx, c_idx, t_idx)] = model.NewBoolVar(
                        f"assegnamento_d{d_idx}_g{g_idx}_c{c_idx}_turno{t_idx}"
                    )

    deficit_turno = {}
    turni_domenicali_effettivi = [nt for nt, quando in dati["quando_turni"].items() if quando == "Domenicale"]

    for g_idx in range(dati["num_giorni"]):
        giorno_str_loop = dati["giorni_mese"][g_idx]
        giorno_settimana_loop = get_giorno_settimana(giorno_str_loop).lower()

        for c_idx in range(dati["num_cantieri"]):
            for nome_turno_loop in dati["turni_per_cantiere"][c_idx]:
                t_idx = dati["nome_turni_map"][nome_turno_loop]
                num_richiesto = dati["dipendenti_per_turno_config"].get(nome_turno_loop, 1)

                is_turno_domenicale = nome_turno_loop in turni_domenicali_effettivi
                is_allowed_combination = False

                if giorno_settimana_loop == "domenica" and is_turno_domenicale:
                    is_allowed_combination = True
                elif giorno_settimana_loop != "domenica" and not is_turno_domenicale:
                    is_allowed_combination = True

                if is_allowed_combination and num_richiesto > 0:
                    deficit_turno[(g_idx, c_idx, t_idx)] = model.NewIntVar(0, num_richiesto, f"deficit_g{g_idx}_c{c_idx}_t{t_idx}")
                    model.Add(deficit_turno[(g_idx, c_idx, t_idx)] == num_richiesto - sum(assegnamenti[(d_loop, g_idx, c_idx, t_idx)] for d_loop in range(dati["num_dipendenti"])))

    print("Variabili del modello create.")
    return assegnamenti, deficit_turno

# MODIFICATA: applica_vincoli_rigidi
def applica_vincoli_rigidi(model, assegnamenti, dati):
    """Applica i vincoli rigidi al modello CP-SAT."""

    # Vincolo Rigido 1: Restrizione Giorno della Settimana / Tipo di Turno ("quando")
    turni_domenicali_effettivi = [nt for nt, quando in dati["quando_turni"].items() if quando == "Domenicale"]
    for g_idx in range(dati["num_giorni"]):
        giorno_str_loop = dati["giorni_mese"][g_idx]
        giorno_settimana_loop = get_giorno_settimana(giorno_str_loop).lower()
        for d_idx in range(dati["num_dipendenti"]):
            for c_idx in range(dati["num_cantieri"]):
                for nome_turno_loop in dati["turni_per_cantiere"][c_idx]:
                    t_idx = dati["nome_turni_map"][nome_turno_loop]
                    is_turno_domenicale = nome_turno_loop in turni_domenicali_effettivi
                    if giorno_settimana_loop == "domenica" and not is_turno_domenicale:
                         model.Add(assegnamenti[(d_idx, g_idx, c_idx, t_idx)] == 0)
                    elif giorno_settimana_loop != "domenica" and is_turno_domenicale:
                         model.Add(assegnamenti[(d_idx, g_idx, c_idx, t_idx)] == 0)

    # Vincolo Rigido 2: Numero MASSIMO di dipendenti per turno e cantiere per giorno
    for g_idx in range(dati["num_giorni"]):
        giorno_str_loop = dati["giorni_mese"][g_idx]
        giorno_settimana_loop = get_giorno_settimana(giorno_str_loop).lower()
        for c_idx in range(dati["num_cantieri"]):
            for nome_turno_loop in dati["turni_per_cantiere"][c_idx]:
                t_idx = dati["nome_turni_map"][nome_turno_loop]
                num_richiesto_max = dati["dipendenti_per_turno_config"].get(nome_turno_loop, 1)
                is_turno_domenicale = nome_turno_loop in turni_domenicali_effettivi
                is_allowed_combination = (giorno_settimana_loop == "domenica" and is_turno_domenicale) or \
                                        (giorno_settimana_loop != "domenica" and not is_turno_domenicale)
                if is_allowed_combination and num_richiesto_max > 0:
                     model.Add(
                         sum(assegnamenti[(d_loop, g_idx, c_idx, t_idx)] for d_loop in range(dati["num_dipendenti"])) <= num_richiesto_max
                     )

    # Vincolo Rigido 3: Certificazioni richieste per i turni
    for g_idx in range(dati["num_giorni"]):
        for c_idx in range(dati["num_cantieri"]):
            for nome_turno_loop in dati["turni_per_cantiere"][c_idx]:
                t_idx = dati["nome_turni_map"][nome_turno_loop]
                certificazione_richiesta = dati["certificazioni_richieste"].get(nome_turno_loop)
                if certificazione_richiesta:
                    for d_idx in range(dati["num_dipendenti"]):
                        dipendente_ha_cert = certificazione_richiesta in dati["dipendenti"][d_idx].get("certificazioni", [])
                        if not dipendente_ha_cert:
                            model.Add(assegnamenti[(d_idx, g_idx, c_idx, t_idx)] == 0)

    # --- NUOVO VINCOLO RIGIDO ---
    # Vincolo Rigido 4: Indisponibilità da file Excel
    print("Applicando vincolo rigido di indisponibilità da Excel...")
    indisponibilita_excel = dati.get("indisponibilita_dipendenti_excel", {}) # Prende il dizionario caricato

    for d_idx in range(dati["num_dipendenti"]):
        # Ottieni l'ID del dipendente dall'indice d_idx
        id_dipendente_attuale = dati["dipendenti"][d_idx]["id"]
        # Ottieni il nome normalizzato del dipendente usando il suo ID
        # dati["nome_dipendente"] ora contiene nomi normalizzati (UPPERCASE)
        nome_dipendente_attuale_norm = dati["nome_dipendente"].get(id_dipendente_attuale)

        if nome_dipendente_attuale_norm is None:
            # Questo non dovrebbe accadere se i dati sono coerenti
            print(f"Attenzione: Nome non trovato per dipendente con indice {d_idx} e ID {id_dipendente_attuale}")
            continue

        if nome_dipendente_attuale_norm in indisponibilita_excel:
            date_non_disponibile_list = indisponibilita_excel[nome_dipendente_attuale_norm]
            # print(f"Debug Excel: Dipendente {nome_dipendente_attuale_norm}, Indisponibile in date: {date_non_disponibile_list}") # Debug
            for g_idx_excel in range(dati["num_giorni"]): # Rinomina g_idx per evitare conflitto di scope
                giorno_str_corrente = dati["giorni_mese"][g_idx_excel] # Formato "YYYY-MM-DD"

                if giorno_str_corrente in date_non_disponibile_list:
                    # Se il dipendente non è disponibile in questo giorno, non può essere assegnato a nessun turno/cantiere
                    # print(f"Vincolo Excel applicato: {nome_dipendente_attuale_norm} non disponibile il {giorno_str_corrente}") # Per debug
                    for c_idx_excel in range(dati["num_cantieri"]): # Rinomina c_idx
                        for nome_turno_loop_excel in dati["turni_per_cantiere"][c_idx_excel]: # Rinomina nome_turno_loop
                            t_idx_excel = dati["nome_turni_map"][nome_turno_loop_excel] # Rinomina t_idx
                            # Aggiungi il vincolo che l'assegnamento deve essere 0
                            model.Add(assegnamenti[(d_idx, g_idx_excel, c_idx_excel, t_idx_excel)] == 0)
        # else:
            # print(f"Debug Excel: Nessuna indisponibilità da Excel per {nome_dipendente_attuale_norm}") # Debug

    print("Vincoli rigidi applicati (inclusa indisponibilità Excel).")


def applica_vincoli_individuali(model, assegnamenti, violazioni, dati):
    """Applica i vincoli individuali dei dipendenti (soft o rigidi) al modello."""

    def applica_singolo_vincolo(dipendente_obj, vincolo_obj, model_inner, assegnamenti_dict_inner, violazioni_dict_inner, dati_inner):
        if vincolo_obj.get("tipo") in ["limite_ore_settimanali", "alternanza_domenicale_globale"]:
            return

        dip_id_val = dipendente_obj["id"]
        dip_idx_val = dati_inner["id_a_indice_dipendente"][dip_id_val]
        tipo_vincolo = vincolo_obj.get("tipo")
        if tipo_vincolo is None: return
        is_soft_val = vincolo_obj.get("soft", False)
        penalita_val = vincolo_obj.get("penalita", 0)

        if tipo_vincolo == "esclusione_temporale":
            vincolo_quando_list = vincolo_obj.get("quando", [])
            vincolo_giorni_nomi_list = vincolo_obj.get("giorno", [])
            if isinstance(vincolo_giorni_nomi_list, str): vincolo_giorni_nomi_list = [vincolo_giorni_nomi_list]
            vincolo_giorni_nomi_list = [g_nome.lower() for g_nome in vincolo_giorni_nomi_list]

            for g_loop_idx, giorno_str_val in enumerate(dati_inner["giorni_mese"]):
                giorno_sett_attuale_val = get_giorno_settimana(giorno_str_val).lower()
                for c_loop_idx in range(dati_inner["num_cantieri"]):
                    for nome_turno_val in dati_inner["turni_per_cantiere"][c_loop_idx]:
                        t_loop_idx = dati_inner["nome_turni_map"][nome_turno_val]
                        turno_quando_attuale_val = dati_inner["quando_turni"].get(nome_turno_val)
                        escludere_questo_turno_flag = False
                        if vincolo_giorni_nomi_list and vincolo_quando_list:
                             if giorno_sett_attuale_val in vincolo_giorni_nomi_list and turno_quando_attuale_val in vincolo_quando_list: escludere_questo_turno_flag = True
                        elif vincolo_giorni_nomi_list:
                            if giorno_sett_attuale_val in vincolo_giorni_nomi_list: escludere_questo_turno_flag = True
                        elif vincolo_quando_list:
                            if turno_quando_attuale_val in vincolo_quando_list: escludere_questo_turno_flag = True
                        elif not vincolo_giorni_nomi_list and not vincolo_quando_list: escludere_questo_turno_flag = False

                        if escludere_questo_turno_flag:
                            if is_soft_val:
                                violazione_var = model_inner.NewBoolVar(f"violazione_escl_temp_d{dip_id_val}_g{g_loop_idx}_c{c_loop_idx}_t{t_loop_idx}")
                                model_inner.Add(assegnamenti_dict_inner[(dip_idx_val, g_loop_idx, c_loop_idx, t_loop_idx)] <= violazione_var)
                                violazioni_dict_inner[("esclusione_temporale", dip_id_val, giorno_str_val, nome_turno_val, dati_inner["nome_cantieri"][c_loop_idx])].append((violazione_var, penalita_val))
                            else:
                                model_inner.Add(assegnamenti_dict_inner[(dip_idx_val, g_loop_idx, c_loop_idx, t_loop_idx)] == 0)

        elif tipo_vincolo == "esclusione_oraria":
            giorno_val = vincolo_obj.get("giorno", "").lower()
            dopo_le_val = vincolo_obj.get("dopo_le")
            fino_a_val = vincolo_obj.get("fino_a")
            for g_loop_idx, giorno_str_val in enumerate(dati_inner["giorni_mese"]):
                giorno_settimana_val = get_giorno_settimana(giorno_str_val).lower()
                if not giorno_val or giorno_settimana_val == giorno_val:
                    for c_loop_idx in range(dati_inner["num_cantieri"]):
                        for nome_turno_val in dati_inner["turni_per_cantiere"][c_loop_idx]:
                            t_loop_idx = dati_inner["nome_turni_map"][nome_turno_val]
                            orario_inizio_val = dati_inner["orario_inizio_turni"].get(nome_turno_val, -1)
                            escludere_per_orario = False
                            if dopo_le_val is not None and orario_inizio_val != -1 and orario_inizio_val >= dopo_le_val: escludere_per_orario = True
                            if fino_a_val is not None and orario_inizio_val != -1 and orario_inizio_val < fino_a_val: escludere_per_orario = True
                            if escludere_per_orario:
                                if is_soft_val:
                                    violazione_var = model_inner.NewBoolVar(f"violazione_esclusione_oraria_d{dip_id_val}_g{g_loop_idx}_c{c_loop_idx}_t{t_loop_idx}")
                                    model_inner.Add(assegnamenti_dict_inner[(dip_idx_val, g_loop_idx, c_loop_idx, t_loop_idx)] <= violazione_var)
                                    violazioni_dict_inner[("esclusione_oraria", dip_id_val, giorno_str_val, nome_turno_val, dati_inner["nome_cantieri"][c_loop_idx])].append((violazione_var, penalita_val))
                                else:
                                    model_inner.Add(assegnamenti_dict_inner[(dip_idx_val, g_loop_idx, c_loop_idx, t_loop_idx)] == 0)

        elif tipo_vincolo == "restrizione_cantieri":
            cantieri_permessi_list = vincolo_obj.get("cantieri", [])
            for g_loop_idx in range(dati_inner["num_giorni"]):
                for c_loop_idx in range(dati_inner["num_cantieri"]):
                    if dati_inner["nome_cantieri"][c_loop_idx] not in cantieri_permessi_list:
                        for nome_turno_val in dati_inner["turni_per_cantiere"][c_loop_idx]:
                            t_loop_idx = dati_inner["nome_turni_map"][nome_turno_val]
                            if is_soft_val:
                                violazione_var = model_inner.NewBoolVar(f"violazione_restrizione_cantieri_d{dip_id_val}_g{g_loop_idx}_c{c_loop_idx}_t{t_loop_idx}")
                                model_inner.Add(assegnamenti_dict_inner[(dip_idx_val, g_loop_idx, c_loop_idx, t_loop_idx)] <= violazione_var)
                                violazioni_dict_inner[("restrizione_cantieri", dip_id_val, dati_inner["giorni_mese"][g_loop_idx], nome_turno_val, dati_inner["nome_cantieri"][c_loop_idx])].append((violazione_var, penalita_val))
                            else:
                                model_inner.Add(assegnamenti_dict_inner[(dip_idx_val, g_loop_idx, c_loop_idx, t_loop_idx)] == 0)

        elif tipo_vincolo == "limite_frequenza":
            quando_val = vincolo_obj.get("quando")
            giorno_val = vincolo_obj.get("giorno", "").lower()
            frequenza_val = vincolo_obj.get("frequenza")
            periodo_val = vincolo_obj.get("periodo")
            if frequenza_val is None: return
            turni_validi_list = [nt for nt in dati_inner["nome_turni_map"] if dati_inner["quando_turni"].get(nt) == quando_val]
            count_var = sum(
                assegnamenti_dict_inner[(dip_idx_val, g_loop_idx, c_loop_idx, dati_inner["nome_turni_map"][nt_val])]
                for g_loop_idx in range(dati_inner["num_giorni"])
                for c_loop_idx in range(dati_inner["num_cantieri"])
                for nt_val in dati_inner["turni_per_cantiere"][c_loop_idx]
                if nt_val in turni_validi_list and (not giorno_val or get_giorno_settimana(dati_inner["giorni_mese"][g_loop_idx]).lower() == giorno_val)
            )
            if is_soft_val:
                max_sforamento_val = len(turni_validi_list) * dati_inner["num_giorni"] * dati_inner["num_cantieri"]
                sforamento_var = model_inner.NewIntVar(0, max_sforamento_val, f"sforamento_lim_freq_d{dip_id_val}_{quando_val}_{giorno_val}")
                model_inner.Add(sforamento_var >= count_var - frequenza_val)
                violazioni_dict_inner[("limite_frequenza", dip_id_val, periodo_val, quando_val, giorno_val)].append((sforamento_var, penalita_val))
            else:
                model_inner.Add(count_var <= frequenza_val)

        elif tipo_vincolo == "dipendenza_turni":
            se_turno_val = vincolo_obj.get("se_turno")
            non_turno_successivo_val = vincolo_obj.get("non_turno_successivo")
            giorni_dopo_val = vincolo_obj.get("giorni_dopo", 1)
            indici_se_turno = [dati_inner["nome_turni_map"][nt] for nt, quando in dati_inner["quando_turni"].items() if quando == se_turno_val]
            indici_non_turno = [dati_inner["nome_turni_map"][nt] for nt, quando in dati_inner["quando_turni"].items() if quando == non_turno_successivo_val]
            if not indici_se_turno or not indici_non_turno: pass
            else:
                 for g_loop_idx in range(dati_inner["num_giorni"] - giorni_dopo_val):
                    somma_se_turno = sum(
                        assegnamenti_dict_inner[(dip_idx_val, g_loop_idx, c_loop_idx, t_loop_idx)]
                        for c_loop_idx in range(dati_inner["num_cantieri"]) for t_loop_idx in indici_se_turno
                        if dati_inner["indice_turni_map"].get(t_loop_idx) in dati_inner["turni_per_cantiere"][c_loop_idx]
                    )
                    somma_non_turno = sum(
                         assegnamenti_dict_inner[(dip_idx_val, g_loop_idx + giorni_dopo_val, c_loop_idx, t_loop_idx)]
                         for c_loop_idx in range(dati_inner["num_cantieri"]) for t_loop_idx in indici_non_turno
                         if dati_inner["indice_turni_map"].get(t_loop_idx) in dati_inner["turni_per_cantiere"][c_loop_idx]
                    )
                    if is_soft_val:
                        violazione_var = model_inner.NewBoolVar(f"violazione_dipendenza_turni_d{dip_id_val}_g{g_loop_idx}_{se_turno_val}_{non_turno_successivo_val}")
                        model_inner.Add(somma_se_turno + somma_non_turno <= 1 + violazione_var)
                        violazioni_dict_inner[("dipendenza_turni", dip_id_val, dati_inner["giorni_mese"][g_loop_idx], se_turno_val, non_turno_successivo_val)].append((violazione_var, penalita_val))
                    else:
                        model_inner.Add(somma_se_turno + somma_non_turno <= 1)

        elif tipo_vincolo == "adiacenza_turni": # Questo sembra un duplicato di limite_turni_giornalieri, verificare se l'intento è diverso
            max_turni_val = vincolo_obj.get("max_turni_giorno", 1)
            for g_loop_idx in range(dati_inner["num_giorni"]):
                count_daily_shifts_var = sum(
                    assegnamenti_dict_inner[(dip_idx_val, g_loop_idx, c_loop_idx, dati_inner["nome_turni_map"][nt_val])]
                    for c_loop_idx in range(dati_inner["num_cantieri"])
                    for nt_val in dati_inner["turni_per_cantiere"][c_loop_idx]
                )
                if is_soft_val:
                    max_daily_shifts_possible = sum(len(dati_inner["turni_per_cantiere"][c_idx]) for c_idx in range(dati_inner["num_cantieri"]))
                    sforamento_var = model_inner.NewIntVar(0, max_daily_shifts_possible, f"sfor_adiacenza_d{dip_id_val}_g{g_loop_idx}")
                    model_inner.Add(sforamento_var >= count_daily_shifts_var - max_turni_val)
                    violazioni_dict_inner[("adiacenza_turni_conteggio", dip_id_val, dati_inner["giorni_mese"][g_loop_idx], "", "")].append((sforamento_var, penalita_val))
                else:
                    model_inner.Add(count_daily_shifts_var <= max_turni_val)

        elif tipo_vincolo == "alternanza_temporale":
            quando1_val = vincolo_obj.get("quando1")
            quando2_val = vincolo_obj.get("quando2")
            settimana_inizio_val = vincolo_obj.get("settimana_inizio", 1)
            # print(f"DEBUG ALTERNANZA INDIVIDUALE: Dipendente {dip_id_val}, Vincolo: {vincolo_obj}")
            for g_loop_idx, giorno_str_val in enumerate(dati_inner["giorni_mese"]):
                settimana_attuale_val = get_settimana_mese(giorno_str_val)
                offset_settimana = settimana_attuale_val - settimana_inizio_val
                turno_da_bloccare_val = quando2_val if offset_settimana % 2 == 0 else quando1_val
                # print(f"DEBUG ALTERNANZA INDIVIDUALE: Data: {giorno_str_val}, Sett.ISO: {settimana_attuale_val}, Offset: {offset_settimana}, Bloccare: {turno_da_bloccare_val}")
                for c_loop_idx in range(dati_inner["num_cantieri"]):
                    for nome_turno_val in dati_inner["turni_per_cantiere"][c_loop_idx]:
                        turno_quando_attuale_val = dati_inner["quando_turni"].get(nome_turno_val)
                        if turno_quando_attuale_val == turno_da_bloccare_val:
                            t_loop_idx = dati_inner["nome_turni_map"][nome_turno_val]
                            if is_soft_val:
                                violazione_var = model_inner.NewBoolVar(f"violazione_alternanza_temp_d{dip_id_val}_g{g_loop_idx}_c{c_loop_idx}_t{t_loop_idx}")
                                model_inner.Add(assegnamenti_dict_inner[(dip_idx_val, g_loop_idx, c_loop_idx, t_loop_idx)] <= violazione_var)
                                violazioni_dict_inner[("alternanza_temporale", dip_id_val, giorno_str_val, nome_turno_val, dati_inner["nome_cantieri"][c_loop_idx])].append((violazione_var, penalita_val))
                            else:
                                # print(f"DEBUG ALTERNANZA INDIVIDUALE RIGIDA: Match! Blocco {dati_inner['nome_dipendente'][dip_id_val]} il {giorno_str_val} Turno: {nome_turno_val}")
                                model_inner.Add(assegnamenti_dict_inner[(dip_idx_val, g_loop_idx, c_loop_idx, t_loop_idx)] == 0)

        elif tipo_vincolo == "limite_turni_giornalieri":
            max_turni_val = vincolo_obj.get("max_turni", 1)
            for g_loop_idx in range(dati_inner["num_giorni"]):
                giorno_str_val = dati_inner["giorni_mese"][g_loop_idx]
                count_turni_giornalieri_var = sum(
                    assegnamenti_dict_inner[(dip_idx_val, g_loop_idx, c_loop_idx, dati_inner["nome_turni_map"][nt_val])]
                    for c_loop_idx in range(dati_inner["num_cantieri"])
                    for nt_val in dati_inner["turni_per_cantiere"][c_loop_idx]
                )
                if is_soft_val:
                    max_daily_shifts_possible = sum(len(dati_inner["turni_per_cantiere"][c_idx]) for c_idx in range(dati_inner["num_cantieri"]))
                    sforamento_var = model_inner.NewIntVar(0, max_daily_shifts_possible, f"sfor_lim_turni_giorn_d{dip_id_val}_g{g_loop_idx}")
                    model_inner.Add(sforamento_var >= count_turni_giornalieri_var - max_turni_val)
                    violazioni_dict_inner[("limite_turni_giornalieri", dip_id_val, giorno_str_val, "", "")].append((sforamento_var, penalita_val))
                else:
                    model_inner.Add(count_turni_giornalieri_var <= max_turni_val)

    for dipendente_item in dati["dipendenti"]:
       for vincolo_item in dipendente_item.get("vincoli", []):
           applica_singolo_vincolo(dipendente_item, vincolo_item, model, assegnamenti, violazioni, dati)
    print("Vincoli individuali applicati.")


def applica_vincolo_limite_ore_settimanali(model, assegnamenti, violazioni, dati):
    """Applica il vincolo (soft o rigido) sul limite di ore settimanali per dipendente."""
    print("Applicando vincolo limite ore settimanali...")
    for dipendente_item in dati["dipendenti"]:
        dip_id_val = dipendente_item["id"]
        dip_idx_val = dati["id_a_indice_dipendente"][dip_id_val]
        limite_ore_vincolo = next((v for v in dipendente_item.get("vincoli", []) if v.get("tipo") == "limite_ore_settimanali"), None)
        is_soft_vincolo_ore = False
        penalita_val_ore = 0
        max_ore_val_float_considerato = dipendente_item.get("max_ore_settimanali", 0)
        if limite_ore_vincolo:
            is_soft_vincolo_ore = limite_ore_vincolo.get("soft", False)
            penalita_val_ore = limite_ore_vincolo.get("penalita", 0)
            max_ore_val_float_considerato = limite_ore_vincolo.get("max_ore", max_ore_val_float_considerato)
        max_ore_val_ch = int(max_ore_val_float_considerato * 100)
        settimane_del_mese = sorted(list(set(get_settimana_mese(g_str) for g_str in dati["giorni_mese"])))
        for settimana_val in settimane_del_mese:
            ore_settimana_var = sum(
                assegnamenti[(dip_idx_val, g_loop_idx, c_loop_idx, dati["nome_turni_map"][nt_val])] * dati["durata_turni"][nt_val]
                for g_loop_idx in range(dati["num_giorni"]) if get_settimana_mese(dati["giorni_mese"][g_loop_idx]) == settimana_val
                for c_loop_idx in range(dati["num_cantieri"]) for nt_val in dati["turni_per_cantiere"][c_loop_idx]
            )
            if max_ore_val_ch > 0:
                if is_soft_vincolo_ore:
                    max_sforamento_settimanale_ch = 50 * 100
                    sforamento_var = model.NewIntVar(0, max_sforamento_settimanale_ch, f"sforamento_ore_settimana_d{dip_id_val}_s{settimana_val}")
                    model.Add(sforamento_var >= ore_settimana_var - max_ore_val_ch)
                    violazioni[("limite_ore_settimanali", dip_id_val, settimana_val, "", "")].append((sforamento_var, penalita_val_ore))
                else:
                     model.Add(ore_settimana_var <= max_ore_val_ch)
    print("Vincolo limite ore settimanali applicato.")


def applica_vincolo_alternanza_domenicale_globale(model, assegnamenti, violazioni, dati):
    """Applica il vincolo (soft o rigido) sull'alternanza domenicale globale."""
    print("Applicando vincolo alternanza domenicale globale...")
    alternanza_config = dati["vincoli_globali"].get("alternanza_domenicale", {})
    if alternanza_config.get("attivo", False):
        settimana_inizio_alternanza = alternanza_config.get("settimana_inizio_alternanza", 1)
        is_soft_alternanza = alternanza_config.get("soft", False)
        penalita_alternanza = alternanza_config.get("penalita", 0)
        turni_domenicali_alternanza_nomi = alternanza_config.get("turni", [])
        for dipendente_item in dati["dipendenti"]:
            dip_id_val = dipendente_item["id"]
            dip_idx_val = dati["id_a_indice_dipendente"][dip_id_val]
            esclusione_domenicale_personale_rigida = any(
                v_item.get("tipo") == "esclusione_temporale" and "Domenicale" in v_item.get("quando", []) and not v_item.get("soft", False)
                for v_item in dipendente_item.get("vincoli", [])
            )
            if esclusione_domenicale_personale_rigida:
                # print(f"DEBUG ALTERNANZA DOM GLOBALE: Dipendente {dip_id_val} ha esclusione domenicale rigida, salto alternanza globale.")
                continue
            for g_loop_idx, giorno_str_val in enumerate(dati["giorni_mese"]):
                if get_giorno_settimana(giorno_str_val).lower() == "domenica":
                    settimana_attuale_val = get_settimana_mese(giorno_str_val)
                    non_dovrebbe_lavorare_questa_domenica = (settimana_attuale_val - settimana_inizio_alternanza) % 2 == 1
                    if non_dovrebbe_lavorare_questa_domenica:
                         for c_loop_idx in range(dati["num_cantieri"]):
                            for nome_turno_val in dati["turni_per_cantiere"][c_loop_idx]:
                                if nome_turno_val in turni_domenicali_alternanza_nomi:
                                    t_loop_idx = dati["nome_turni_map"][nome_turno_val]
                                    if is_soft_alternanza:
                                        violazione_var = model.NewBoolVar(f"violazione_alternanza_globale_d{dip_id_val}_g{g_loop_idx}_c{c_loop_idx}_t{t_loop_idx}")
                                        model.Add(assegnamenti[(dip_idx_val, g_loop_idx, c_loop_idx, t_loop_idx)] <= violazione_var)
                                        violazioni[("alternanza_domenicale_globale", dip_id_val, giorno_str_val, nome_turno_val, dati["nome_cantieri"][c_loop_idx])].append((violazione_var, penalita_alternanza))
                                    else:
                                        model.Add(assegnamenti[(dip_idx_val, g_loop_idx, c_loop_idx, t_loop_idx)] == 0)
    print("Vincolo alternanza domenicale globale applicato.")


def definisci_obiettivo(model, assegnamenti, violazioni, deficit_turno, dati, config):
    """Definisce la funzione obiettivo per il modello CP-SAT."""
    print("Definendo la funzione obiettivo...")
    priorita_dipendenti = {d["id"]: calcola_priorita_vincoli(d) for d in dati["dipendenti"]}
    tot_turni_per_dip_cantieri_list = []
    max_turni_possibili_val = sum(len(dati["turni_per_cantiere"][c_idx]) for c_idx in range(dati["num_cantieri"])) * dati["num_giorni"]
    for d_idx in range(dati["num_dipendenti"]):
        tot_turni_dip_var = sum(
            assegnamenti[(d_idx, g_idx, c_idx, dati["nome_turni_map"][nt_val])]
            for g_idx in range(dati["num_giorni"]) for c_idx in range(dati["num_cantieri"])
            for nt_val in dati["turni_per_cantiere"][c_idx]
        )
        tot_turni_per_dip_cantieri_list.append(tot_turni_dip_var)
    somma_turni_assegnati_var = sum(tot_turni_per_dip_cantieri_list)
    varianza_expr = 0
    if dati["num_dipendenti"] > 0:
        media_turni_cantieri_val = somma_turni_assegnati_var / dati["num_dipendenti"] if isinstance(somma_turni_assegnati_var, (int, float)) else max_turni_possibili_val / dati["num_dipendenti"]
        diff_quadratiche_cantieri_list = []
        for d_idx in range(dati["num_dipendenti"]):
            diff_expr = tot_turni_per_dip_cantieri_list[d_idx] - int(media_turni_cantieri_val)
            diff_var_for_quad = model.NewIntVar(-max_turni_possibili_val, max_turni_possibili_val, f"diff_var_for_quad_{d_idx}")
            model.Add(diff_var_for_quad == diff_expr)
            max_diff_quad_val = min(max_turni_possibili_val**2, (1 << 31) -1 )
            diff_quad_var = model.NewIntVar(0, max_diff_quad_val if max_diff_quad_val >=0 else (1 << 31) -1 , f"diff_quad_cantiere_{d_idx}")
            model.AddMultiplicationEquality(diff_quad_var, [diff_var_for_quad, diff_var_for_quad])
            diff_quadratiche_cantieri_list.append(diff_quad_var)
        varianza_expr = sum(diff_quadratiche_cantieri_list) if diff_quadratiche_cantieri_list else 0
    penalita_violazioni_tot_expr = 0
    for chiave_viol, lista_viol_obj in violazioni.items():
        for violazione_var, penalita_val in lista_viol_obj:
             penalita_violazioni_tot_expr += violazione_var * penalita_val
    somma_deficit_domenicali_expr = 0
    somma_deficit_altri_giorni_expr = 0
    turni_domenicali_effettivi = [nt for nt, quando in dati["quando_turni"].items() if quando == "Domenicale"]
    for (g_idx, c_idx, t_idx), deficit_var in deficit_turno.items():
        nome_turno_loop = dati["indice_turni_map"][t_idx]
        if nome_turno_loop in turni_domenicali_effettivi:
            somma_deficit_domenicali_expr += deficit_var
        else:
            somma_deficit_altri_giorni_expr += deficit_var
    peso_varianza_val = config.get("pesi_obiettivo", {}).get("varianza", 1)
    peso_turni_val = config.get("pesi_obiettivo", {}).get("turni_assegnati", 1000)
    peso_violazioni_val = config.get("pesi_obiettivo", {}).get("violazioni_soft", 10)
    peso_priorita_val = config.get("pesi_obiettivo", {}).get("priorita_dipendenti", 500)
    peso_deficit_turno_domenicale_val = config.get("pesi_obiettivo", {}).get("deficit_domenicale", 2500)
    peso_deficit_turno_altri_giorni_val = config.get("pesi_obiettivo", {}).get("deficit_altri_giorni", 1500)
    somma_turni_pesata_expr = sum(
        tot_turni_per_dip_cantieri_list[d_idx] * priorita_dipendenti[dati["dipendenti"][d_idx]["id"]]
        for d_idx in range(dati["num_dipendenti"])
    )
    if dati["num_dipendenti"] > 0:
        obiettivo_expr = (
            peso_turni_val * somma_turni_assegnati_var +
            peso_priorita_val * somma_turni_pesata_expr -
            peso_varianza_val * varianza_expr -
            peso_violazioni_val * penalita_violazioni_tot_expr -
            peso_deficit_turno_domenicale_val * somma_deficit_domenicali_expr -
            peso_deficit_turno_altri_giorni_val * somma_deficit_altri_giorni_expr
        )
        model.Maximize(obiettivo_expr)
        print("Funzione obiettivo definita.")
        return obiettivo_expr, penalita_violazioni_tot_expr, somma_deficit_domenicali_expr, somma_deficit_altri_giorni_expr
    else:
        print("Nessun dipendente configurato. Obiettivo non definito.")
        return None, None, None, None


def risolvi_modello(model, tempo_limite_secondi=60.0):
    """Risolve il modello CP-SAT."""
    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = tempo_limite_secondi
    solver.parameters.log_search_progress = True
    solver.parameters.num_search_workers = 8
    print(f"Avvio risolutore CP-SAT con limite di {tempo_limite_secondi} secondi...")
    status = solver.Solve(model)
    print(f"Risoluzione completata. Stato: {solver.StatusName(status)}")
    return status, solver

def raccogli_risultati(solver, status, assegnamenti, deficit_turno, violazioni, dati):
    """Raccoglie i risultati del solver e li organizza per i report."""
    print("Raccogliendo i risultati del solver...")
    assegnamenti_effettuati_count = 0
    turni_non_completamente_assegnati_list = []
    assegnamenti_per_excel = []
    piano_per_dipendente_pivot = defaultdict(lambda: defaultdict(list))
    turni_per_dipendente_count = defaultdict(int)
    ore_per_dipendente_count = defaultdict(float)
    weekly_hours_assigned = defaultdict(lambda: defaultdict(float))
    violazioni_rilevate_list = [] # Spostata qui per essere accessibile anche se non ci sono soluzioni

    if status == cp_model.OPTIMAL or status == cp_model.FEASIBLE:
        num_assegnamenti_richiesti_totali = 0
        turni_domenicali_effettivi = [nt for nt, quando in dati["quando_turni"].items() if quando == "Domenicale"]
        for g_idx in range(dati["num_giorni"]):
            giorno_str_loop = dati["giorni_mese"][g_idx]
            giorno_settimana_loop = get_giorno_settimana(giorno_str_loop).lower()
            for c_idx in range(dati["num_cantieri"]):
                for nome_turno_loop in dati["turni_per_cantiere"][c_idx]:
                     is_turno_domenicale = nome_turno_loop in turni_domenicali_effettivi
                     is_allowed_combination = (giorno_settimana_loop == "domenica" and is_turno_domenicale) or \
                                              (giorno_settimana_loop != "domenica" and not is_turno_domenicale)
                     if is_allowed_combination:
                         num_assegnamenti_richiesti_totali += dati["dipendenti_per_turno_config"].get(nome_turno_loop, 1)

        for d_idx in range(dati["num_dipendenti"]):
            # Usa l'ID del dipendente per ottenere il nome normalizzato
            dip_id_attuale = dati["dipendenti"][d_idx]["id"]
            dip_nome = dati["nome_dipendente"].get(dip_id_attuale, f"ID SCONOSCIUTO {dip_id_attuale}")

            for g_idx in range(dati["num_giorni"]):
                giorno_str_loop = dati["giorni_mese"][g_idx]
                giorno_sett_loop = get_giorno_settimana(giorno_str_loop)
                settimana_attuale = get_settimana_mese(giorno_str_loop)
                for c_idx in range(dati["num_cantieri"]):
                    nome_cantiere_loop = dati["nome_cantieri"][c_idx]
                    for nome_turno_loop in dati["turni_per_cantiere"][c_idx]:
                        t_idx = dati["nome_turni_map"].get(nome_turno_loop)
                        if t_idx is None: continue
                        if (d_idx, g_idx, c_idx, t_idx) in assegnamenti and solver.Value(assegnamenti[(d_idx, g_idx, c_idx, t_idx)]) == 1:
                             assegnamenti_effettuati_count += 1
                             assegnamenti_per_excel.append({
                                "data": giorno_str_loop, "giorno_settimana": giorno_sett_loop,
                                "turno": nome_turno_loop, "cantiere": nome_cantiere_loop,
                                "dipendente": dip_nome, "violazioni_dipendente": ""
                             })
                             piano_per_dipendente_pivot[dip_nome][giorno_str_loop].append(f"{nome_turno_loop} ({nome_cantiere_loop})")
                             turni_per_dipendente_count[dip_nome] += 1
                             durata_turno_ore = dati["durata_turni"].get(nome_turno_loop, 0) / 100.0
                             ore_per_dipendente_count[dip_nome] += durata_turno_ore
                             weekly_hours_assigned[dip_nome][settimana_attuale] += durata_turno_ore

        turni_domenicali_effettivi = [nt for nt, quando in dati["quando_turni"].items() if quando == "Domenicale"]
        for g_idx in range(dati["num_giorni"]):
             giorno_str_loop = dati["giorni_mese"][g_idx]
             giorno_sett_loop = get_giorno_settimana(giorno_str_loop)
             giorno_settimana_loop = get_giorno_settimana(giorno_str_loop).lower()
             for c_idx in range(dati["num_cantieri"]):
                 nome_cantiere_loop = dati["nome_cantieri"][c_idx]
                 for nome_turno_loop in dati["turni_per_cantiere"][c_idx]:
                    t_idx = dati["nome_turni_map"].get(nome_turno_loop)
                    if t_idx is None: continue
                    num_richiesto_per_turno = dati["dipendenti_per_turno_config"].get(nome_turno_loop, 1)
                    num_assegnati_al_turno = sum(
                        1 for d_loop_idx in range(dati["num_dipendenti"])
                        if (d_loop_idx, g_idx, c_idx, t_idx) in assegnamenti and solver.Value(assegnamenti[(d_loop_idx, g_idx, c_idx, t_idx)]) == 1
                    )
                    is_turno_domenicale = nome_turno_loop in turni_domenicali_effettivi
                    is_allowed_combination = (giorno_settimana_loop == "domenica" and is_turno_domenicale) or \
                                             (giorno_settimana_loop != "domenica" and not is_turno_domenicale)
                    if is_allowed_combination and num_richiesto_per_turno > 0 and num_assegnati_al_turno < num_richiesto_per_turno:
                        turni_non_completamente_assegnati_list.append(
                            f"{giorno_str_loop}, {nome_turno_loop}, {nome_cantiere_loop} "
                            f"(Richiesti: {num_richiesto_per_turno}, Assegnati: {num_assegnati_al_turno})"
                        )
                        for _ in range(num_richiesto_per_turno - num_assegnati_al_turno):
                             assegnamenti_per_excel.append({
                                "data": giorno_str_loop, "giorno_settimana": giorno_sett_loop,
                                "turno": nome_turno_loop, "cantiere": nome_cantiere_loop,
                                "dipendente": "NON ASSEGNATO", "violazioni_dipendente": ""
                             })
        assegnamenti_per_excel.sort(key=lambda x: (x["data"], x["cantiere"], x["turno"], x["dipendente"]))

        # print("\n--- DEBUG RILEVAMENTO VIOLAZIONI SOFT (RACCOGLI RISULTATI) ---")
        for chiave_viol, lista_viol_obj in violazioni.items():
            tipo_viol = chiave_viol[0]; dip_id_viol = chiave_viol[1]; info_extra = chiave_viol[2:]
            valore_penalita_accumulata = 0.0; conteggio_violazioni_istanza = 0
            # if tipo_viol == "alternanza_temporale": print(f"DEBUG RILEVAMENTO VIOLAZIONI (RR): Key {chiave_viol}")
            for i, (violazione_var, penalita_val) in enumerate(lista_viol_obj):
                try:
                     if isinstance(violazione_var, (cp_model.IntVar, cp_model.BoolVar)):
                        violazione_valore = solver.Value(violazione_var)
                        # if tipo_viol == "alternanza_temporale": print(f"DEBUG RILEVAMENTO VIOLAZIONI (RR):   Var {violazione_var.Name()}, Value: {violazione_valore}, Penalty: {penalita_val}")
                        if violazione_valore > 0:
                            valore_penalita_accumulata += violazione_valore * penalita_val
                            conteggio_violazioni_istanza += violazione_valore
                except Exception: pass # Ignora se la variabile non è valutabile
            # if tipo_viol == "alternanza_temporale": print(f"DEBUG RILEVAMENTO VIOLAZIONI (RR): Key {chiave_viol}, Acc.Penalty: {valore_penalita_accumulata}, Count: {conteggio_violazioni_istanza}")
            if valore_penalita_accumulata > 0:
                dipendente_nome_viol = dati["nome_dipendente"].get(dip_id_viol, f"ID {dip_id_viol}")
                descrizione_viol = f"Dipendente {dipendente_nome_viol} violato {tipo_viol}. Conteggio/Sforamento: {conteggio_violazioni_istanza}, Penalità totale: {valore_penalita_accumulata:.1f}."
                descrizione_viol += " Dettagli: "
                if tipo_viol in ["esclusione_temporale", "esclusione_oraria", "restrizione_cantieri", "alternanza_temporale", "alternanza_domenicale_globale"]:
                     descrizione_viol += f"Giorno: {info_extra[0] if len(info_extra) > 0 else 'N/D'}, Turno: {info_extra[1] if len(info_extra) > 1 else 'N/D'}, Cantiere: {info_extra[2] if len(info_extra) > 2 else 'N/D'}"
                elif tipo_viol == "limite_frequenza": descrizione_viol += f"Periodo: {info_extra[0]}, Quando: {info_extra[1]}, Giorno: {info_extra[2]}"
                elif tipo_viol == "dipendenza_turni": descrizione_viol += f"Giorno inizio: {info_extra[0]}, Se Turno: {info_extra[1]}, Non Turno Successivo: {info_extra[2]}"
                elif tipo_viol in ["adiacenza_turni_conteggio", "limite_turni_giornalieri"]: descrizione_viol += f"Giorno: {info_extra[0]}"
                elif tipo_viol == "limite_ore_settimanali": descrizione_viol += f"Settimana: {info_extra[0]} (ISO)"
                else: descrizione_viol += f"Info extra: {info_extra}"
                violazioni_rilevate_list.append(descrizione_viol)
        # print("--- FINE DEBUG RILEVAMENTO VIOLAZIONI SOFT (RACCOGLI RISULTATI) ---")

        for item_excel in assegnamenti_per_excel:
            if item_excel["dipendente"] != "NON ASSEGNATO":
                 dip_nome_current = item_excel["dipendente"]
                 violazioni_specifiche_assegnamento = []
                 for v_desc in violazioni_rilevate_list:
                      if dip_nome_current in v_desc and item_excel["data"] in v_desc:
                          match_details = True
                          if "Turno:" in v_desc and item_excel["turno"] not in v_desc: match_details = False
                          if "Cantiere:" in v_desc and item_excel["cantiere"] not in v_desc: match_details = False
                          if match_details:
                             summary_part = v_desc.split(". Dettagli:")[0] if ". Dettagli:" in v_desc else v_desc
                             violazioni_specifiche_assegnamento.append(summary_part)
                 item_excel["violazioni_dipendente"] = "; ".join(violazioni_specifiche_assegnamento)
    else: # Se non ottimale o fattibile
        num_assegnamenti_richiesti_totali = 0 # Non calcolabile senza soluzione
        print("Nessuna soluzione trovata dal solver. Impossibile calcolare assegnamenti o violazioni.")


    print("Risultati del solver raccolti.")
    return {
        "assegnamenti_effettuati_count": assegnamenti_effettuati_count,
        "num_assegnamenti_richiesti_totali": num_assegnamenti_richiesti_totali,
        "turni_non_completamente_assegnati_list": turni_non_completamente_assegnati_list,
        "assegnamenti_per_excel": assegnamenti_per_excel,
        "piano_per_dipendente_pivot": piano_per_dipendente_pivot,
        "turni_per_dipendente_count": turni_per_dipendente_count,
        "ore_per_dipendente_count": ore_per_dipendente_count,
        "weekly_hours_assigned": weekly_hours_assigned,
        "violazioni_rilevate_list": violazioni_rilevate_list,
        "nome_dipendente": dati["nome_dipendente"]
    }

def crea_report_excel(risultati, dati, anno_pianificazione, mese_pianificazione):
    """Crea e salva il report Excel unico con più fogli."""
    print("Creando report Excel...")
    script_dir = os.path.dirname(os.path.abspath(__file__))
    base_name_excel = f"piano_turni_{anno_pianificazione}_{mese_pianificazione}"
    extension_excel = ".xlsx"
    directory = script_dir
    version_excel = get_next_version_number(base_name_excel, extension_excel, directory)
    excel_file_unico = os.path.join(directory, f"{base_name_excel}_v{version_excel}{extension_excel}")
    workbook_unico = Workbook()
    sheet_dettaglio = workbook_unico.active
    sheet_dettaglio.title = "Piano Turni Dettaglio"
    intestazione_dettaglio = ["Data", "Giorno", "Turno", "Cantiere", "Dipendente Assegnato", "Violazioni Dipendente (Riepilogo)"]
    sheet_dettaglio.append(intestazione_dettaglio)
    for item_excel in risultati["assegnamenti_per_excel"]:
        data_obj_excel = datetime.strptime(item_excel["data"], "%Y-%m-%d")
        data_formattata_excel = data_obj_excel.strftime("%d/%m/%Y")
        riga_dati_excel = [
            data_formattata_excel, item_excel["giorno_settimana"], item_excel["turno"],
            item_excel["cantiere"], item_excel["dipendente"], item_excel["violazioni_dipendente"]
        ]
        sheet_dettaglio.append(riga_dati_excel)
    for column_cells in sheet_dettaglio.columns:
        max_length = 0; column = column_cells[0].column_letter
        for cell in column_cells:
            if cell.value is not None:
                cell_str = str(cell.value); lines = cell_str.splitlines()
                if lines: max_length = max(max_length, len(lines[0]))
        adjusted_length = min(max_length, 100) if column == 'F' else max_length
        sheet_dettaglio.column_dimensions[column].width = adjusted_length + 2

    if risultati["violazioni_rilevate_list"]:
        sheet_violazioni = workbook_unico.create_sheet("Riepilogo Violazioni Soft")
        sheet_violazioni.append(["Descrizione Violazione Soft Dettagliata"])
        for v_text in risultati["violazioni_rilevate_list"]: sheet_violazioni.append([v_text])
        if sheet_violazioni["A"]:
             max_length_viol = max((len(str(cell.value)) if cell.value is not None else 0) for cell in sheet_violazioni["A"])
             sheet_violazioni.column_dimensions['A'].width = min(max_length_viol + 2, 150)

    if risultati["turni_non_completamente_assegnati_list"]:
        sheet_non_coperti = workbook_unico.create_sheet("Turni Non Coperti")
        sheet_non_coperti.append(["Descrizione Turno Non Coperto"])
        for tnc_text in risultati["turni_non_completamente_assegnati_list"]: sheet_non_coperti.append([tnc_text])
        if sheet_non_coperti["A"]:
             max_length_tnc = max((len(str(cell.value)) if cell.value is not None else 0) for cell in sheet_non_coperti["A"])
             sheet_non_coperti.column_dimensions['A'].width = min(max_length_tnc + 2, 150)

    sheet_pivot = workbook_unico.create_sheet("Piano Turni Pivot Dipendenti")
    date_riga_pivot = ["Dipendente"] + [datetime.strptime(g_str, "%Y-%m-%d").strftime("%d/%m/%Y") for g_str in dati["giorni_mese"]]
    sheet_pivot.append(date_riga_pivot)
    giorni_riga_pivot = [""] + [get_giorno_settimana(g_str) for g_str in dati["giorni_mese"]]
    sheet_pivot.append(giorni_riga_pivot)
    tutti_i_nomi_dipendenti_ordinati = sorted(dati["nome_dipendente"].values())
    for dipendente_nome_pivot_loop in tutti_i_nomi_dipendenti_ordinati:
        riga_pivot = [dipendente_nome_pivot_loop]
        for giorno_str_loop in dati["giorni_mese"]:
            turni_list_pivot = risultati["piano_per_dipendente_pivot"].get(dipendente_nome_pivot_loop, {}).get(giorno_str_loop, [])
            cella_pivot = "\n".join(sorted(turni_list_pivot)) if turni_list_pivot else ""
            riga_pivot.append(cella_pivot)
        sheet_pivot.append(riga_pivot)
    for column_cells in sheet_pivot.columns:
        max_length = 0;
        if not column_cells: continue
        column = column_cells[0].column_letter
        for cell in column_cells:
            if cell.value is not None:
                cell_str = str(cell.value); lines = cell_str.splitlines()
                if lines: max_length = max(max_length, len(lines[0]))
        sheet_pivot.column_dimensions[column].width = max_length + 2

    sheet_summary_operator = workbook_unico.create_sheet("Riepilogo Turni e Ore")
    summary_operator_headers = ["Dipendente", "Turni Totali", "Ore Totali"]
    settimane_del_mese = sorted(list(set(get_settimana_mese(g_str) for g_str in dati["giorni_mese"])))
    summary_operator_headers.extend([f"Sett. {week_num} Ore" for week_num in settimane_del_mese])
    sheet_summary_operator.append(summary_operator_headers)
    for dip_nome in tutti_i_nomi_dipendenti_ordinati:
        total_shifts = risultati["turni_per_dipendente_count"].get(dip_nome, 0)
        total_hours = risultati["ore_per_dipendente_count"].get(dip_nome, 0.0)
        row_data = [dip_nome, total_shifts, total_hours]
        for week_num in settimane_del_mese:
            weekly_hours = risultati["weekly_hours_assigned"].get(dip_nome, {}).get(week_num, 0.0)
            row_data.append(weekly_hours)
        sheet_summary_operator.append(row_data)
    for column_cells in sheet_summary_operator.columns:
        max_length = 0
        if not column_cells: continue
        column = column_cells[0].column_letter
        for cell in column_cells:
            if cell.value is not None:
                cell_str = str(cell.value); lines = cell_str.splitlines()
                if lines: max_length = max(max_length, len(lines[0]))
        sheet_summary_operator.column_dimensions[column].width = max_length + 2
    try:
        workbook_unico.save(excel_file_unico)
        print(f"\n✅ Piano turni unico salvato in '{excel_file_unico}'")
    except Exception as e:
        print(f"\n❌ Errore nel salvataggio del file Excel unico: {e}")
    print("Report Excel creato.")


def stampa_riepiloghi_console(risultati, solver, status, obiettivo_expr, penalita_violazioni_tot_expr, somma_deficit_domenicali_expr, somma_deficit_altri_giorni_expr, dati):
    """Stampa i riepiloghi dei risultati nella console."""
    print("\n--- Riepilogo Risoluzione ---")
    print(f"Stato solver: {solver.StatusName(status)}")
    if obiettivo_expr is not None and (status == cp_model.OPTIMAL or status == cp_model.FEASIBLE):
        try: print(f"Valore obiettivo: {solver.Value(obiettivo_expr)}")
        except Exception as e: print(f"Impossibile ottenere il valore obiettivo: {e}")
    if penalita_violazioni_tot_expr is not None and (status == cp_model.OPTIMAL or status == cp_model.FEASIBLE):
         try: print(f"Penalità totali violazioni soft: {solver.Value(penalita_violazioni_tot_expr)}")
         except Exception as e: print(f"Impossibile ottenere il valore delle penalità totali soft: {e}")
    if somma_deficit_domenicali_expr is not None and (status == cp_model.OPTIMAL or status == cp_model.FEASIBLE):
         try: print(f"Totale deficit slot turni DOMENICALI: {solver.Value(somma_deficit_domenicali_expr)}")
         except Exception as e: print(f"Impossibile ottenere il valore del deficit domenicale: {e}")
    if somma_deficit_altri_giorni_expr is not None and (status == cp_model.OPTIMAL or status == cp_model.FEASIBLE):
         try: print(f"Totale deficit slot turni ALTRI GIORNI: {solver.Value(somma_deficit_altri_giorni_expr)}")
         except Exception as e: print(f"Impossibile ottenere il valore del deficit altri giorni: {e}")

    # Gestione del caso in cui num_assegnamenti_richiesti_totali potrebbe essere 0 o non definito
    if "num_assegnamenti_richiesti_totali" in risultati and risultati["num_assegnamenti_richiesti_totali"] > 0:
        percentuale_completamento = (risultati["assegnamenti_effettuati_count"] / risultati["num_assegnamenti_richiesti_totali"]) * 100
        print(f"\nPercentuale di completamento (slot dipendente): {percentuale_completamento:.2f}% ({risultati['assegnamenti_effettuati_count']}/{risultati['num_assegnamenti_richiesti_totali']})")
    elif "assegnamenti_effettuati_count" in risultati : # Se ci sono assegnamenti ma non richieste (improbabile ma per sicurezza)
        print(f"\nAssegnamenti effettuati: {risultati['assegnamenti_effettuati_count']} (richieste totali non definite o zero)")
    else:
        print("\nDati di completamento non disponibili (nessuna soluzione o richieste zero).")


    if risultati["violazioni_rilevate_list"]:
        print("\nViolazioni Soft Rilevate (sommario per tipo/dipendente):")
        for v_desc in risultati["violazioni_rilevate_list"]: print(v_desc)
    if risultati["turni_non_completamente_assegnati_list"]:
        print("\nTurni Non Completamente Assegnati:")
        for tna_desc in risultati["turni_non_completamente_assegnati_list"]: print(tna_desc)
    print("\nDistribuzione dei turni per operatore:")
    tutti_i_nomi_dipendenti_ordinati = sorted(dati["nome_dipendente"].values())
    for dip_nome in tutti_i_nomi_dipendenti_ordinati:
        count_val = risultati["turni_per_dipendente_count"].get(dip_nome, 0)
        ore_val = risultati["ore_per_dipendente_count"].get(dip_nome, 0.0)
        print(f"{dip_nome}: {count_val} turni, {ore_val:.2f} ore")
    print("\n--- Fine Riepilogo ---")


# MODIFICATA: main
def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    config_path = os.path.join(script_dir, "config2.json")
    # Definisci il percorso del file Excel per le indisponibilità
    # Assicurati che questo file esista nella stessa directory dello script o fornisci il percorso completo.
    file_indisponibilita_path = os.path.join(script_dir, "indisponibilita_dipendenti.xlsx")

    config = carica_configurazione(config_path)
    # Passa il percorso del file di indisponibilità a inizializza_dati
    dati = inizializza_dati(config, file_indisponibilita_path)

    model_cantieri = cp_model.CpModel()
    violazioni = defaultdict(list)
    assegnamenti, deficit_turno = crea_variabili_modello(model_cantieri, dati)

    # Applica i vincoli rigidi (ora include la lettura da Excel)
    applica_vincoli_rigidi(model_cantieri, assegnamenti, dati)

    applica_vincoli_individuali(model_cantieri, assegnamenti, violazioni, dati)
    applica_vincolo_limite_ore_settimanali(model_cantieri, assegnamenti, violazioni, dati)
    applica_vincolo_alternanza_domenicale_globale(model_cantieri, assegnamenti, violazioni, dati)
    obiettivo_expr, penalita_violazioni_tot_expr, somma_deficit_domenicali_expr, somma_deficit_altri_giorni_expr = definisci_obiettivo(model_cantieri, assegnamenti, violazioni, deficit_turno, dati, config)
    
    tempo_limite_risoluzione = config.get("parametri_risolutore", {}).get("tempo_limite_secondi", 60.0)
    status_cantieri, solver_cantieri = risolvi_modello(model_cantieri, tempo_limite_risoluzione)

    if status_cantieri == cp_model.OPTIMAL or status_cantieri == cp_model.FEASIBLE:
        risultati = raccogli_risultati(solver_cantieri, status_cantieri, assegnamenti, deficit_turno, violazioni, dati)
        stampa_riepiloghi_console(risultati, solver_cantieri, status_cantieri, obiettivo_expr, penalita_violazioni_tot_expr, somma_deficit_domenicali_expr, somma_deficit_altri_giorni_expr, dati)
        crea_report_excel(risultati, dati, config["anno"], config["mese"])
    else:
        print("Nessuna soluzione trovata o il modello non è fattibile.")
        # Per un debug più approfondito in caso di infattibilità, potresti voler stampare statistiche o log del solver.
        # Ad esempio, se il solver lo supporta, potresti cercare di ottenere informazioni sui vincoli in conflitto.
        # print(f"Statistiche del solver: {solver_cantieri.ResponseStats()}") # Utile per il debug avanzato

if __name__ == "__main__":
    main()
